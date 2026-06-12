"""
generador.py
Genera un ZIP en memoria con un Excel por cuenta seleccionada.
- Días anteriores a hoy → /statements (trae saldo)
- Hoy                  → /movements/dia (saldo calculado desde el día anterior)
"""

import io
import os
import zipfile
import importlib
import http.client
import json
from datetime import date, datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from auth import obtener_token

# ─────────────────────────────────────────────
# CONFIG EMPRESAS
# ─────────────────────────────────────────────

EMPRESAS = {
    "eliantus": {
        "cuentas": "srcELIANTUS.CodigoBancosEliantus",
    },
    "elementa": {
        "cuentas": "srcELEMENTA.CodigoBancosElementa",
    },
    "integra": {
        "cuentas": "srcINTEGRA.CodigoBancosINTEGRA",
    },
}

# ─────────────────────────────────────────────
# ESTILOS EXCEL
# ─────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(bold=True, color="FFFFFF")
SALDO_FILL    = PatternFill("solid", fgColor="D6E4F0")
SALDO_FONT    = Font(bold=True, color="1F4E79")
SUBTOTAL_FILL = PatternFill("solid", fgColor="EBF5FB")
SUBTOTAL_FONT = Font(bold=True)
ACCOUNT_FILL  = PatternFill("solid", fgColor="FFF2CC")
ACCOUNT_FONT  = Font(bold=True, color="7B6000")
HOY_FILL      = PatternFill("solid", fgColor="E2EFDA")   # verde suave para fila resumen de hoy
HOY_FONT      = Font(bold=True, color="375623")
NUMBER_FORMAT = '#,##0.00'


def _aplicar_fila(ws, row_num, fill, font):
    for cell in ws[row_num]:
        cell.fill = fill
        cell.font = font


# ─────────────────────────────────────────────
# LLAMADAS A LA API
# ─────────────────────────────────────────────

def _get_statements(cuenta, token, customer_id, client_id, desde, hasta):
    """
    /v1/statements → días anteriores.
    Devuelve lista de statements, cada uno con movement_detail, opening_balance, etc.
    """
    conn = http.client.HTTPSConnection("api-gw.interbanking.com.ar")
    headers = {
        "client_id":     client_id,
        "Authorization": f"Bearer {token}",
        "accept":        "application/json",
    }
    path = (
        f"/api/prod/v1/accounts/{cuenta.numero}/statements"
        f"?account-type={cuenta.tipo}"
        f"&bank-number={cuenta.banco}"
        f"&currency={cuenta.peso}"
        f"&customer-id={customer_id}"
        f"&date-since={desde}"
        f"&date-until={hasta}"
        f"&limit=10000"
    )
    conn.request("GET", path, headers=headers)
    data = json.loads(conn.getresponse().read().decode("utf-8"))
    return data.get("general_data", {}), data.get("statements", [])


def _get_movimientos_dia(cuenta, token, customer_id, client_id):
    """
    /v2/movements/dia → movimientos de hoy.
    No trae saldo — se calcula a partir del ending_balance del día anterior.
    """
    conn = http.client.HTTPSConnection("api-gw.interbanking.com.ar")
    headers = {
        "client_id":     client_id,
        "Authorization": f"Bearer {token}",
        "accept":        "application/json",
    }
    path = (
        f"/api/prod/v2/accounts/{cuenta.numero}/movements/dia"
        f"?account-type={cuenta.tipo}"
        f"&bank-number={cuenta.banco}"
        f"&currency={cuenta.peso}"
        f"&customer-id={customer_id}"
        f"&limit=10000"
        f"&page=0"
    )
    conn.request("GET", path, headers=headers)
    data = json.loads(conn.getresponse().read().decode("utf-8"))
    return data.get("general_data", {}), data.get("movements_detail", [])


# ─────────────────────────────────────────────
# ESCRITURA EXCEL
# ─────────────────────────────────────────────

def _fmt_fecha(fecha_iso):
    """2026-05-21 → 21/05/2026"""
    anio, mes, dia = fecha_iso[:10].split("-")
    return f"{dia}/{mes}/{anio}"


def _escribir_encabezado_cuenta(ws, cuenta, desde, hasta):
    ws.append([f"▶ {cuenta.nombre}  ({desde} → {hasta})",
               None, None, None, None, None, None, None, None])
    _aplicar_fila(ws, ws.max_row, ACCOUNT_FILL, ACCOUNT_FONT)

    ws.append(["Fecha", "Importe", "Tipo", "CUIT", "Descripción",
               "Saldo Inicial", "Saldo Final", "Total Débitos", "Total Créditos"])
    for cell in ws[ws.max_row]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _escribir_statement(ws, statement):
    """Escribe las filas de un statement (día con saldo conocido)."""
    fecha_fmt  = _fmt_fecha(statement["operation_date"])
    opening    = statement.get("opening_balance") or 0
    ending     = statement.get("ending_balance")
    deb_total  = statement.get("debits_total_amount")
    cred_total = statement.get("credits_total_amount")
    movimientos = statement.get("movement_detail", [])

    if not movimientos:
        ws.append([fecha_fmt, None, "RESUMEN DIA", None, "(sin movimientos)",
                   opening, ending, deb_total, cred_total])
        _aplicar_fila(ws, ws.max_row, SUBTOTAL_FILL, SUBTOTAL_FONT)
        return

    saldo = opening
    for mov in sorted(movimientos, key=lambda x: x.get("process_date", "")):
        importe   = mov.get("amount") or 0
        saldo_ini = saldo
        saldo_fin = saldo + importe
        saldo     = saldo_fin

        descripcion = " | ".join(filter(None, [
            mov.get("code_description_bank"),
            mov.get("code_description_ib"),
            mov.get("code_description_standard"),
        ]))

        ws.append([fecha_fmt, importe, mov.get("debit_credit_type"),
                   mov.get("customer_cuit"), descripcion,
                   saldo_ini, saldo_fin, None, None])

    # Fila resumen del día
    ws.append([fecha_fmt, None, "RESUMEN DIA", None, None,
               opening, ending, deb_total, cred_total])
    _aplicar_fila(ws, ws.max_row, SALDO_FILL, SALDO_FONT)


def _escribir_movimientos_dia(ws, movimientos, saldo_inicial_hoy):
    """
    Escribe los movimientos de hoy.
    saldo_inicial_hoy = ending_balance del último statement (día anterior).
    Si no hay statements previos, saldo_inicial_hoy = None → no mostramos columnas saldo.
    """
    hoy       = date.today().isoformat()
    fecha_fmt = _fmt_fecha(hoy)

    if not movimientos:
        ws.append([fecha_fmt, None, "SIN MOVIMIENTOS HOY", None, None,
                   saldo_inicial_hoy, None, None, None])
        _aplicar_fila(ws, ws.max_row, SUBTOTAL_FILL, SUBTOTAL_FONT)
        return

    saldo = saldo_inicial_hoy if saldo_inicial_hoy is not None else 0

    debitos  = 0.0
    creditos = 0.0

    for mov in sorted(movimientos, key=lambda x: x.get("process_date", "")):
        importe   = mov.get("amount") or 0
        saldo_ini = saldo
        saldo_fin = saldo + importe
        saldo     = saldo_fin

        if importe < 0:
            debitos  += importe
        else:
            creditos += importe

        descripcion = " | ".join(filter(None, [
            mov.get("code_description_bank"),
            mov.get("code_description_ib"),
            mov.get("code_description_standard"),
        ]))

        ws.append([fecha_fmt, importe, mov.get("debit_credit_type"),
                   mov.get("customer_cuit"), descripcion,
                   saldo_ini, saldo_fin, None, None])

    # Fila resumen hoy (saldo final calculado)
    ws.append([fecha_fmt, None, "RESUMEN HOY ✓", None, "(saldo calculado)",
               saldo_inicial_hoy, saldo, debitos, creditos])
    _aplicar_fila(ws, ws.max_row, HOY_FILL, HOY_FONT)


def _aplicar_formato_hoja(ws):
    col_numericas = [2, 6, 7, 8, 9]
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column in col_numericas and isinstance(cell.value, (int, float)):
                cell.number_format = NUMBER_FORMAT

    for col_letra, ancho in {
        "A": 13, "B": 16, "C": 16, "D": 16,
        "E": 45, "F": 17, "G": 17, "H": 17, "I": 17,
    }.items():
        ws.column_dimensions[col_letra].width = ancho

    ws.freeze_panes = "A3"


def _nombre_archivo(cuenta, desde, hasta):
    def fmt(f): 
        a, m, d = f.split("-")
        return f"{d}-{m}"
    return f"{cuenta.abreviatura}_{fmt(desde)}_{fmt(hasta)}.xlsx"


# ─────────────────────────────────────────────
# GENERAR EXCEL POR CUENTA
# ─────────────────────────────────────────────

def _generar_excel_cuenta(cuenta, token, customer_id, client_id, desde, hasta):
    hoy = date.today().isoformat()

    wb = Workbook()
    ws = wb.active
    ws.title = "Extracto"

    _escribir_encabezado_cuenta(ws, cuenta, desde, hasta)

    saldo_final_ayer = None   # se alimenta con el ending_balance del último statement

    # ── STATEMENTS (días anteriores a hoy) ──────────────────────────────
    # Si hasta == hoy pedimos statements hasta ayer; si hasta < hoy, pedimos todo.
    hasta_statements = hasta
    if hasta >= hoy:
        ayer = (date.today() - timedelta(days=1)).isoformat()
        hasta_statements = ayer

    if desde <= hasta_statements:
        _, statements = _get_statements(
            cuenta, token, customer_id, client_id, desde, hasta_statements
        )
        statements_sorted = sorted(statements, key=lambda x: x["operation_date"])

        for statement in statements_sorted:
            _escribir_statement(ws, statement)
            # Guardamos el ending_balance del último día para usarlo como saldo inicial de hoy
            if statement.get("ending_balance") is not None:
                saldo_final_ayer = statement["ending_balance"]

    # ── MOVIMIENTOS DE HOY ───────────────────────────────────────────────
    if hasta >= hoy:
        _, movimientos_hoy = _get_movimientos_dia(
            cuenta, token, customer_id, client_id
        )
        _escribir_movimientos_dia(ws, movimientos_hoy, saldo_final_ayer)

    _aplicar_formato_hoja(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────
# FUNCIÓN PRINCIPAL
# ─────────────────────────────────────────────

def generar_zip(empresa: str, desde: str, hasta: str, cuentas_seleccionadas=None):
    """
    Genera un ZIP en memoria con un Excel por cuenta.
    Retorna (zip_bytes, resultados).
    """
    config  = EMPRESAS[empresa]
    mod     = importlib.import_module(config["cuentas"])
    CUENTAS = mod.CUENTAS

    token, customer_id = obtener_token(empresa)
    client_id          = os.environ.get(f"{empresa.upper()}_CLIENTID", "")

    cuentas_a_usar = cuentas_seleccionadas if cuentas_seleccionadas else CUENTAS
    resultados     = []

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for cuenta in cuentas_a_usar:
            try:
                excel_bytes = _generar_excel_cuenta(
                    cuenta, token, customer_id, client_id, desde, hasta
                )
                filename = _nombre_archivo(cuenta, desde, hasta)
                zf.writestr(filename, excel_bytes)
                resultados.append((cuenta, True))
                print(f"[OK] {cuenta.nombre} → {filename}")

            except Exception as e:
                print(f"[ERROR] {cuenta.nombre}: {e}")
                resultados.append((cuenta, False))

    zip_buf.seek(0)
    return zip_buf.read(), resultados