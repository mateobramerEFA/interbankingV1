"""
generador_dia.py
Genera un ZIP en memoria con un Excel por cuenta seleccionada (movimientos del día).
Mismo formato visual que generador.py.
"""

import io
import os
import zipfile
import http.client
import json
import importlib
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from auth import obtener_token

# ─────────────────────────────────────────────
# CONFIG EMPRESAS
# ─────────────────────────────────────────────

EMPRESAS = {
    "eliantus": { "cuentas": "srcELIANTUS.CodigoBancosEliantus" },
    "elementa": { "cuentas": "srcELEMENTA.CodigoBancosElementa" },
    "integra":  { "cuentas": "srcINTEGRA.CodigoBancosINTEGRA"  },
}

# ─────────────────────────────────────────────
# ESTILOS
# ─────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(bold=True, color="FFFFFF")
SALDO_FILL    = PatternFill("solid", fgColor="D6E4F0")
SALDO_FONT    = Font(bold=True, color="1F4E79")
SUBTOTAL_FILL = PatternFill("solid", fgColor="EBF5FB")
SUBTOTAL_FONT = Font(bold=True)
ACCOUNT_FILL  = PatternFill("solid", fgColor="FFF2CC")
ACCOUNT_FONT  = Font(bold=True, color="7B6000")
NUMBER_FORMAT = '#,##0.00'


def _aplicar_fila(ws, row_num, fill, font):
    for cell in ws[row_num]:
        cell.fill = fill
        cell.font = font


# ─────────────────────────────────────────────
# LLAMADA A LA API
# ─────────────────────────────────────────────

def _obtener_movimientos_dia(cuenta, token, customer_id, client_id):
    hoy  = date.today().isoformat()

    conn = http.client.HTTPSConnection("api-gw.interbanking.com.ar")
    headers = {
        "client_id":     client_id,
        "Authorization": f"Bearer {token}",
        "accept":        "application/json",
    }

    path = (
        f"/api/prod/v2/accounts/{cuenta.numero}/movements/dia"
        f"?account-type={cuenta.tipo}"
        f"&bank-number={cuenta.banco}"
        f"&currency={cuenta.peso}"
        f"&customer-id={customer_id}"
        f"&limit=10000"
        f"&page=0"
    )

    conn.request("GET", path, headers=headers)
    data = json.loads(conn.getresponse().read().decode("utf-8"))
    return data.get("general_data", {}), data.get("movements_detail", [])


# ─────────────────────────────────────────────
# GENERAR EXCEL PARA UNA CUENTA
# ─────────────────────────────────────────────

def _generar_excel(cuenta, general_data, movimientos):
    hoy = date.today().isoformat()
    anio, mes, dia = hoy.split("-")
    fecha_fmt = f"{dia}/{mes}/{anio}"

    wb = Workbook()
    ws = wb.active
    ws.title = "Mov. del día"

    ws.append([f"▶ {cuenta.nombre}  — Movimientos del día {fecha_fmt}",
               None, None, None, None, None, None, None, None])
    _aplicar_fila(ws, ws.max_row, ACCOUNT_FILL, ACCOUNT_FONT)

    ws.append(["Fecha", "Importe", "Tipo", "CUIT", "Descripción",
               "Saldo Inicial", "Saldo Final", "Total Débitos", "Total Créditos"])
    for cell in ws[ws.max_row]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    if not movimientos:
        ws.append([fecha_fmt, None, "SIN MOVIMIENTOS", None, None,
                   None, None, None, None])
        _aplicar_fila(ws, ws.max_row, SUBTOTAL_FILL, SUBTOTAL_FONT)
    else:
        opening    = general_data.get("opening_balance") or 0
        ending     = general_data.get("ending_balance")
        deb_total  = general_data.get("debits_total_amount")
        cred_total = general_data.get("credits_total_amount")

        saldo_corriente = opening

        for mov in sorted(movimientos, key=lambda x: x.get("process_date", "")):
            importe   = mov.get("amount") or 0
            saldo_ini = saldo_corriente
            saldo_fin = saldo_corriente + importe
            saldo_corriente = saldo_fin

            descripcion = " | ".join(filter(None, [
                mov.get("code_description_bank"),
                mov.get("code_description_ib"),
                mov.get("code_description_standard"),
            ]))

            ws.append([fecha_fmt, importe, mov.get("debit_credit_type"),
                       mov.get("customer_cuit"), descripcion,
                       saldo_ini, saldo_fin, None, None])

        ws.append([fecha_fmt, None, "RESUMEN DIA", None, None,
                   opening, ending, deb_total, cred_total])
        _aplicar_fila(ws, ws.max_row, SALDO_FILL, SALDO_FONT)

    col_numericas = [2, 6, 7, 8, 9]
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            if cell.column in col_numericas and isinstance(cell.value, (int, float)):
                cell.number_format = NUMBER_FORMAT

    for col_letra, ancho in {
        "A": 13, "B": 16, "C": 16, "D": 16,
        "E": 45, "F": 17, "G": 17, "H": 17, "I": 17,
    }.items():
        ws.column_dimensions[col_letra].width = ancho

    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────
# FUNCIÓN PRINCIPAL
# ─────────────────────────────────────────────

def generar_zip_dia(empresa: str, indices: list):
    """
    Genera un ZIP en memoria con un Excel por cuenta seleccionada.
    Retorna (zip_bytes, resultados).
    """
    config  = EMPRESAS[empresa]
    mod     = importlib.import_module(config["cuentas"])
    cuentas = mod.CUENTAS

    token, customer_id = obtener_token(empresa)
    client_id          = os.environ.get(f"{empresa.upper()}_CLIENTID", "")

    hoy        = date.today().isoformat()
    resultados = []
    zip_buf    = io.BytesIO()

    with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for i in indices:
            if i < 0 or i >= len(cuentas):
                continue
            cuenta = cuentas[i]
            try:
                general_data, movimientos = _obtener_movimientos_dia(
                    cuenta, token, customer_id, client_id
                )
                excel_bytes = _generar_excel(cuenta, general_data, movimientos)
                filename    = f"{cuenta.abreviatura}_DIA_{hoy}.xlsx"
                zf.writestr(filename, excel_bytes)
                resultados.append((cuenta.abreviatura, bool(movimientos)))
            except Exception as e:
                print(f"[ERROR] {cuenta.nombre}: {e}")
                resultados.append((cuenta.abreviatura, False))

    zip_buf.seek(0)
    return zip_buf.read(), resultados